#!/usr/bin/env -S uv run --script
"""HTTP protocol plugin for fetching data from HTTP/HTTPS endpoints."""
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "requests>=2.31.0",
#   "openpyxl>=3.1.0",
# ]
# [tool.jn]
# matches = [
#   "^https?://.*"
# ]
# ///

import json
import io
import sys
from typing import Iterator
from urllib.parse import urlparse

import requests


# Format detection mapping
FORMAT_DETECT = {
    # Content types
    "application/json": "json",
    "application/x-ndjson": "ndjson",
    "text/csv": "csv",
    # Common Excel content types
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xlsx",
    # File extensions
    ".json": "json",
    ".jsonl": "ndjson",
    ".csv": "csv",
    ".tsv": "csv",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
}


def error_record(error_type: str, message: str, **extra) -> dict:
    """Create standardized error record."""
    return {"_error": True, "type": error_type, "message": message, **extra}


def reads(
    url: str,
    method: str = "GET",
    headers: dict = None,
    auth: tuple = None,
    timeout: int = 30,
    verify_ssl: bool = True,
    force_format: str = None,
) -> Iterator[dict]:
    """Fetch data from HTTP/HTTPS URL and yield NDJSON records.

    Args:
        url: The URL to fetch (required)
        method: HTTP method (default: 'GET')
        headers: Dict of HTTP headers
        auth: Tuple of (username, password) for Basic auth
        timeout: Request timeout in seconds (default: 30)
        verify_ssl: Verify SSL certificates (default: True)
        force_format: Force specific format ('json', 'csv', 'ndjson', 'text')

    Yields:
        Dict records from the response, or error records
    """
    headers = headers or {}

    # Read request body from stdin for POST/PUT/PATCH
    data = None
    if method.upper() in ("POST", "PUT", "PATCH"):
        data = sys.stdin.read()

    # Make request with streaming
    response = requests.request(
        method,
        url,
        headers=headers,
        auth=auth,
        data=data,
        timeout=timeout,
        verify=verify_ssl,
        stream=True,
    )

    if not response.ok:
        yield error_record(
            "http_error",
            f"HTTP {response.status_code}: {response.reason}",
            url=url,
            status_code=response.status_code,
        )
        return

    # Detect format
    content_type = response.headers.get("Content-Type", "")
    file_ext = "." + urlparse(url).path.split(".")[-1].lower()

    # Determine format using lookup dict
    if force_format:
        fmt = force_format
    else:
        # Try content-type first
        fmt = None
        for key, detected_fmt in FORMAT_DETECT.items():
            if key in content_type:
                fmt = detected_fmt
                break
        # Fall back to file extension
        if not fmt:
            fmt = FORMAT_DETECT.get(file_ext, "text")

    # Format handlers dict
    handlers = {
        "json": lambda: _parse_json(response, url),
        "ndjson": lambda: _parse_ndjson(response),
        "csv": lambda: [
            {"content": response.text, "content_type": "text/csv", "url": url}
        ],
        "xlsx": lambda: _parse_xlsx(response),
        "text": lambda: [
            {"content": response.text, "content_type": content_type, "url": url}
        ],
    }

    # Dispatch to handler
    handler = handlers.get(fmt, handlers["text"])
    yield from handler()


def _parse_json(response: requests.Response, url: str) -> Iterator[dict]:
    """Parse JSON response and yield records."""
    try:
        data = response.json()
    except json.JSONDecodeError as e:
        yield error_record("json_decode_error", str(e), url=url)
        return

    if isinstance(data, list):
        for item in data:
            yield item if isinstance(item, dict) else {"value": item}
    elif isinstance(data, dict):
        yield data
    else:
        yield {"value": data}


def _parse_ndjson(response: requests.Response) -> Iterator[dict]:
    """Parse NDJSON response line by line."""
    for line in response.iter_lines(decode_unicode=True):
        line = line.strip()
        if not line:
            continue
        try:
            yield json.loads(line)
        except json.JSONDecodeError as e:
            yield error_record("ndjson_decode_error", str(e), line=line[:100])


def _parse_xlsx(response: requests.Response) -> Iterator[dict]:
    """Parse Excel (XLSX/XLSM) response body into NDJSON records."""
    # Import locally to avoid dependency when not needed
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover - dynamic environment
        yield error_record("missing_dependency", f"openpyxl not available: {e}")
        return

    # Read binary content
    data = response.content

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        yield error_record("xlsx_load_error", str(e))
        return

    # Use first sheet by default
    sheet = wb[wb.sheetnames[0]]

    rows = iter(sheet.iter_rows(values_only=True))
    header = next(rows, None)
    if not header:
        wb.close()
        return

    header = [str(col) if col is not None else f"Column_{i+1}" for i, col in enumerate(header)]

    for row in rows:
        if all(cell is None for cell in row):
            continue
        record = {}
        for i, col_name in enumerate(header):
            value = row[i] if i < len(row) else None
            if value is not None and hasattr(value, "isoformat"):
                value = value.isoformat()
            record[col_name] = value
        yield record

    wb.close()


if __name__ == "__main__":
    import argparse
    import os

    parser = argparse.ArgumentParser(description="HTTP protocol plugin")
    parser.add_argument("--mode", choices=["read"], help="Operation mode")
    parser.add_argument("url", nargs="?", help="URL to fetch")
    parser.add_argument(
        "--method",
        default="GET",
        choices=["GET", "POST", "PUT", "PATCH", "DELETE"],
        help="HTTP method",
    )
    parser.add_argument(
        "--headers",
        type=json.loads,
        default={},
        help="HTTP headers as JSON",
    )
    parser.add_argument("--auth", help="Basic auth as 'username:password'")
    parser.add_argument("--timeout", type=int, default=30, help="Request timeout")
    parser.add_argument(
        "--no-verify-ssl", dest="verify_ssl", action="store_false", help="Disable SSL verification"
    )
    parser.add_argument(
        "--format", choices=["json", "ndjson", "csv", "text"], help="Force format"
    )

    args = parser.parse_args()

    if not args.mode or not args.url:
        parser.error("--mode and URL are required")

    # Parse auth
    auth = None
    if args.auth:
        if ":" not in args.auth:
            parser.error("--auth must be 'username:password'")
        auth = tuple(args.auth.split(":", 1))

    # Substitute env vars in headers
    headers = {}
    for key, value in args.headers.items():
        if isinstance(value, str) and value.startswith("${") and value.endswith("}"):
            env_var = value[2:-1]
            env_value = os.environ.get(env_var)
            if not env_value:
                print(
                    json.dumps(
                        error_record("env_var_not_set", f"Environment variable {env_var} not set")
                    ),
                    flush=True,
                )
                sys.exit(1)
            headers[key] = env_value
        else:
            headers[key] = value

    # Call reads() with direct args (no config dict)
    try:
        for record in reads(
            url=args.url,
            method=args.method,
            headers=headers,
            auth=auth,
            timeout=args.timeout,
            verify_ssl=args.verify_ssl,
            force_format=args.format,
        ):
            print(json.dumps(record), flush=True)
    except requests.exceptions.RequestException as e:
        # Errors are data: emit an error record and exit successfully
        print(
            json.dumps(error_record("request_exception", str(e), url=args.url)),
            flush=True,
        )
        sys.exit(0)
