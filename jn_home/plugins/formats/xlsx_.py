#!/usr/bin/env -S uv run --script
"""Parse XLSX (Excel) files and convert to NDJSON.

Supports multiple parsing modes:
- simple (default): Treat Excel like CSV, first row is headers
- stats: Inspect workbook structure (sheets, dimensions, merged cells)
- raw: Output every cell with full metadata
- table: Extract specific region with explicit configuration
"""
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1.0"
# ]
# [tool.jn]
# matches = [
#   ".*\\.xlsx$",
#   ".*\\.xlsm$"
# ]
# ///

import io
import json
import sys
from typing import Iterator, Optional


def col_letter(col_num: int) -> str:
    """Convert 1-based column number to Excel letter (1=A, 27=AA)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def col_number(col_letter: str) -> int:
    """Convert Excel column letter to 1-based number (A=1, AA=27)."""
    result = 0
    for char in col_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def parse_range(range_str: str) -> dict:
    """Parse Excel range like 'A1:D10' or 'Sheet1!A1:D10'.

    Returns dict with: sheet (optional), min_row, max_row, min_col, max_col
    """
    import re

    sheet = None
    if "!" in range_str:
        sheet, range_str = range_str.split("!", 1)

    # Handle column-only ranges like A:D
    col_match = re.match(r"^([A-Z]+):([A-Z]+)$", range_str, re.I)
    if col_match:
        return {
            "sheet": sheet,
            "min_col": col_number(col_match.group(1)),
            "max_col": col_number(col_match.group(2)),
            "min_row": None,
            "max_row": None,
        }

    # Handle row-only ranges like 1:5
    row_match = re.match(r"^(\d+):(\d+)$", range_str)
    if row_match:
        return {
            "sheet": sheet,
            "min_row": int(row_match.group(1)),
            "max_row": int(row_match.group(2)),
            "min_col": None,
            "max_col": None,
        }

    # Handle cell ranges like A1:D10 or single cell A1
    range_match = re.match(r"^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", range_str, re.I)
    if range_match:
        min_col = col_number(range_match.group(1))
        min_row = int(range_match.group(2))
        if range_match.group(3):
            max_col = col_number(range_match.group(3))
            max_row = int(range_match.group(4))
        else:
            max_col = min_col
            max_row = min_row
        return {
            "sheet": sheet,
            "min_col": min_col,
            "max_col": max_col,
            "min_row": min_row,
            "max_row": max_row,
        }

    raise ValueError(f"Invalid range: {range_str}")


def cell_in_range(row: int, col: int, range_info: dict) -> bool:
    """Check if cell is within range."""
    min_row = range_info.get("min_row")
    max_row = range_info.get("max_row")
    min_col = range_info.get("min_col")
    max_col = range_info.get("max_col")

    if min_row is not None and row < min_row:
        return False
    if max_row is not None and row > max_row:
        return False
    if min_col is not None and col < min_col:
        return False
    if max_col is not None and col > max_col:
        return False
    return True


def get_cell_type(cell) -> str:
    """Get cell type code: s=string, n=number, f=formula, b=boolean, d=date."""
    if cell.data_type == 'f':
        return 'f'
    if cell.data_type == 'b':
        return 'b'
    if cell.data_type == 'd':
        return 'd'
    if cell.value is None:
        return 'n'  # null treated as number
    if isinstance(cell.value, bool):
        return 'b'
    if isinstance(cell.value, (int, float)):
        return 'n'
    return 's'


def get_merge_range(cell, merged_ranges) -> tuple:
    """Check if cell is in a merged range. Returns (range_str, is_origin)."""
    for merged in merged_ranges:
        if cell.coordinate in merged:
            range_str = str(merged)
            is_origin = cell.coordinate == merged.start_cell.coordinate
            return range_str, is_origin
    return None, None


def reads_simple(config: dict) -> Iterator[dict]:
    """Simple mode: Treat Excel like CSV. Fast path for well-structured files."""
    import openpyxl

    sheet_selector = config.get("sheet", 0)
    skip_rows = config.get("skip_rows", 0)

    xlsx_data = sys.stdin.buffer.read()
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)

    # Select sheet
    if isinstance(sheet_selector, int):
        sheet_names = workbook.sheetnames
        if sheet_selector < 0 or sheet_selector >= len(sheet_names):
            raise ValueError(f"Sheet index {sheet_selector} out of range (0-{len(sheet_names)-1})")
        sheet = workbook[sheet_names[sheet_selector]]
    else:
        if sheet_selector not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_selector}' not found. Available: {workbook.sheetnames}")
        sheet = workbook[sheet_selector]

    rows = iter(sheet.iter_rows(values_only=True))

    # Skip initial rows
    for _ in range(skip_rows):
        next(rows, None)

    # Read header row
    header = next(rows, None)
    if not header:
        return

    # Convert header to strings, handle empty cells
    header = [str(col) if col is not None else f"Column_{i+1}" for i, col in enumerate(header)]

    # Yield data rows
    for row in rows:
        if all(cell is None for cell in row):
            continue

        record = {}
        for i, col_name in enumerate(header):
            value = row[i] if i < len(row) else None
            if value is not None and hasattr(value, 'isoformat'):
                value = value.isoformat()
            record[col_name] = value

        yield record

    workbook.close()


def reads_stats(config: dict) -> Iterator[dict]:
    """Stats mode: Inspect workbook structure."""
    import openpyxl

    xlsx_data = sys.stdin.buffer.read()
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=False)

    for idx, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        # Get dimensions
        dimensions = sheet.dimensions or ""
        min_row = sheet.min_row or 0
        max_row = sheet.max_row or 0
        min_col = sheet.min_column or 0
        max_col = sheet.max_column or 0

        # Get merged ranges
        merged_ranges = [str(r) for r in sheet.merged_cells.ranges]

        # Get named tables
        named_tables = []
        if hasattr(sheet, 'tables'):
            for table in sheet.tables.values():
                named_tables.append({
                    "name": table.name,
                    "range": table.ref
                })

        # Get first row values
        first_row = []
        if max_row >= 1:
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col)
                first_row.append(cell.value)

        # Get first column values
        first_col = []
        if max_col >= 1:
            for row in range(1, min(max_row + 1, 20)):  # Limit to 20 rows
                cell = sheet.cell(row=row, column=1)
                first_col.append(cell.value)

        yield {
            "sheet": sheet_name,
            "index": idx,
            "dimensions": dimensions,
            "rows": max_row,
            "cols": max_col,
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
            "merged_ranges": merged_ranges,
            "named_tables": named_tables,
            "first_row": first_row,
            "first_col": first_col,
        }

    workbook.close()


def reads_raw(config: dict) -> Iterator[dict]:
    """Raw mode: Output every non-empty cell with full metadata."""
    import openpyxl

    sheet_selector = config.get("sheet")  # None = all sheets
    range_str = config.get("range")
    formulas_mode = config.get("formulas", "computed")

    xlsx_data = sys.stdin.buffer.read()

    # Load workbook - need two loads for formulas
    if formulas_mode == "text":
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=False)
        workbook_data = None
    elif formulas_mode == "both":
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=False)
        workbook_data = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=True)
    else:  # computed (default)
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=True)
        workbook_data = None

    # Parse range filter
    range_info = None
    range_sheet = None
    if range_str:
        range_info = parse_range(range_str)
        range_sheet = range_info.get("sheet")

    # Determine sheets to process
    if sheet_selector is not None:
        if isinstance(sheet_selector, int):
            sheet_names = [workbook.sheetnames[sheet_selector]]
        else:
            sheet_names = [sheet_selector]
    elif range_sheet:
        sheet_names = [range_sheet]
    else:
        sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_data = workbook_data[sheet_name] if workbook_data else None

        merged_ranges = list(sheet.merged_cells.ranges)

        # Get hidden rows and columns
        hidden_rows = set()
        hidden_cols = set()
        for row_num, rd in sheet.row_dimensions.items():
            if rd.hidden:
                hidden_rows.add(row_num)
        for col_letter_key, cd in sheet.column_dimensions.items():
            if cd.hidden:
                # Convert column letter to number
                hidden_cols.add(col_number(col_letter_key))

        for row in sheet.iter_rows():
            for cell in row:
                # Skip empty cells
                if cell.value is None and cell.data_type != 'f':
                    # Check if cell is part of a merge
                    merge_range, is_origin = get_merge_range(cell, merged_ranges)
                    if not merge_range:
                        continue

                row_num = cell.row
                col_num = cell.column

                # Apply range filter
                if range_info and not cell_in_range(row_num, col_num, range_info):
                    continue

                ref = f"{col_letter(col_num)}{row_num}"
                cell_type = get_cell_type(cell)

                record = {
                    "sheet": sheet_name,
                    "row": row_num,
                    "col": col_num,
                    "ref": ref,
                    "value": cell.value,
                    "type": cell_type,
                }

                # Add computed value for formulas
                if cell_type == 'f' and sheet_data:
                    data_cell = sheet_data.cell(row=row_num, column=col_num)
                    record["computed"] = data_cell.value

                # Add merge info
                merge_range, is_origin = get_merge_range(cell, merged_ranges)
                if merge_range:
                    record["merge"] = merge_range
                    record["merge_origin"] = is_origin

                # Add hidden flag
                if row_num in hidden_rows or col_num in hidden_cols:
                    record["hidden"] = True

                # Add comment
                if cell.comment:
                    record["comment"] = cell.comment.text

                # Add number format (skip General)
                if cell.number_format and cell.number_format != "General":
                    record["format"] = cell.number_format

                # Convert datetime values
                if hasattr(record["value"], 'isoformat'):
                    record["value"] = record["value"].isoformat()

                yield record

    workbook.close()
    if workbook_data:
        workbook_data.close()


def reads_table(config: dict) -> Iterator[dict]:
    """Table mode: Extract specific region as a table with explicit configuration."""
    import openpyxl

    sheet_selector = config.get("sheet", 0)
    range_str = config.get("range")
    header_row = config.get("header_row", 1)  # 1-indexed within range
    header_col = config.get("header_col")  # Column letter for transposed data
    merge_strategy = config.get("merge_strategy", "origin")
    skip_empty = config.get("skip_empty", True)
    formulas_mode = config.get("formulas", "computed")

    xlsx_data = sys.stdin.buffer.read()

    # Load workbook
    if formulas_mode == "text":
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=False)
    else:
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=True)

    # Select sheet
    if isinstance(sheet_selector, int):
        sheet_names = workbook.sheetnames
        if sheet_selector < 0 or sheet_selector >= len(sheet_names):
            raise ValueError(f"Sheet index {sheet_selector} out of range (0-{len(sheet_names)-1})")
        sheet = workbook[sheet_names[sheet_selector]]
    else:
        if sheet_selector not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_selector}' not found. Available: {workbook.sheetnames}")
        sheet = workbook[sheet_selector]

    # Parse range
    if range_str:
        range_info = parse_range(range_str)
        min_row = range_info.get("min_row", 1)
        max_row = range_info.get("max_row", sheet.max_row)
        min_col = range_info.get("min_col", 1)
        max_col = range_info.get("max_col", sheet.max_column)
    else:
        min_row = 1
        max_row = sheet.max_row
        min_col = 1
        max_col = sheet.max_column

    merged_ranges = list(sheet.merged_cells.ranges)

    def get_cell_value(row, col):
        """Get cell value, handling merged cells according to strategy."""
        cell = sheet.cell(row=row, column=col)
        value = cell.value

        # Check if in merged range
        for merged in merged_ranges:
            if cell.coordinate in merged:
                if merge_strategy == "origin":
                    # Return value only from origin, None for others
                    if cell.coordinate != merged.start_cell.coordinate:
                        value = None
                    else:
                        value = sheet.cell(row=merged.min_row, column=merged.min_col).value
                elif merge_strategy == "fill":
                    # Fill all cells with origin value
                    value = sheet.cell(row=merged.min_row, column=merged.min_col).value
                break

        # Convert datetime
        if value is not None and hasattr(value, 'isoformat'):
            value = value.isoformat()

        return value

    if header_col:
        # Transposed data: headers in column, records in rows
        header_col_num = col_number(header_col)

        # Read headers from column
        headers = []
        for row in range(min_row, max_row + 1):
            val = get_cell_value(row, header_col_num)
            headers.append(str(val) if val is not None else f"Row_{row}")

        # Each column after header_col is a record
        for col in range(min_col, max_col + 1):
            if col == header_col_num:
                continue

            record = {}
            all_empty = True
            for row_idx, row in enumerate(range(min_row, max_row + 1)):
                value = get_cell_value(row, col)
                if value is not None:
                    all_empty = False
                record[headers[row_idx]] = value

            if skip_empty and all_empty:
                continue

            yield record
    else:
        # Normal data: headers in row
        header_row_abs = min_row + header_row - 1

        # Read headers
        headers = []
        for col in range(min_col, max_col + 1):
            val = get_cell_value(header_row_abs, col)
            headers.append(str(val) if val is not None else f"Column_{col - min_col + 1}")

        # Read data rows
        for row in range(header_row_abs + 1, max_row + 1):
            record = {}
            all_empty = True

            for col_idx, col in enumerate(range(min_col, max_col + 1)):
                value = get_cell_value(row, col)
                if value is not None:
                    all_empty = False
                record[headers[col_idx]] = value

            if skip_empty and all_empty:
                continue

            yield record

    workbook.close()


def writes(config: Optional[dict] = None) -> None:
    """Read NDJSON from stdin, write XLSX to stdout."""
    import openpyxl

    config = config or {}
    sheet_name = config.get("sheet", "Sheet1")

    records = []
    for line in sys.stdin:
        line = line.strip()
        if line:
            records.append(json.loads(line))

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if not records:
        output = io.BytesIO()
        workbook.save(output)
        sys.stdout.buffer.write(output.getvalue())
        return

    # Get all unique keys
    all_keys = []
    seen = set()
    for record in records:
        for key in record:
            if key not in seen:
                all_keys.append(key)
                seen.add(key)

    # Write header row
    for col_idx, key in enumerate(all_keys, start=1):
        sheet.cell(row=1, column=col_idx, value=key)

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            value = record.get(key)
            sheet.cell(row=row_idx, column=col_idx, value=value)

    output = io.BytesIO()
    workbook.save(output)
    sys.stdout.buffer.write(output.getvalue())


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="XLSX format plugin - read/write Excel files with multiple parsing modes"
    )
    parser.add_argument(
        "--mode",
        choices=["read", "simple", "stats", "raw", "table", "write"],
        help="Operation mode: simple/read (CSV-like), stats (metadata), raw (cell-by-cell), table (region), write (NDJSON to xlsx)",
    )
    parser.add_argument(
        "--sheet",
        help="Sheet name or index (default: 0 for first sheet, all for raw mode)",
    )
    parser.add_argument(
        "--skip-rows",
        type=int,
        default=0,
        help="Rows to skip before header (simple mode)",
    )
    parser.add_argument(
        "--range",
        help="Cell range in A1 notation (raw/table mode)",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="Header row number within range (table mode, 1-indexed)",
    )
    parser.add_argument(
        "--header-col",
        help="Header column for transposed data (table mode)",
    )
    parser.add_argument(
        "--merge-strategy",
        choices=["origin", "fill", "expand"],
        default="origin",
        help="How to handle merged cells (table mode)",
    )
    parser.add_argument(
        "--skip-empty",
        type=lambda x: x.lower() == 'true',
        default=True,
        help="Skip completely empty rows (table mode)",
    )
    parser.add_argument(
        "--formulas",
        choices=["computed", "text", "both"],
        default="computed",
        help="Formula output mode",
    )

    args = parser.parse_args()

    if not args.mode:
        parser.error("--mode is required")

    # Build config
    config = {}

    if args.sheet:
        try:
            config["sheet"] = int(args.sheet)
        except ValueError:
            config["sheet"] = args.sheet

    if args.skip_rows:
        config["skip_rows"] = args.skip_rows

    if args.range:
        config["range"] = args.range

    if args.header_row:
        config["header_row"] = args.header_row

    if args.header_col:
        config["header_col"] = args.header_col

    if args.merge_strategy:
        config["merge_strategy"] = args.merge_strategy

    config["skip_empty"] = args.skip_empty
    config["formulas"] = args.formulas

    # Route to appropriate mode
    if args.mode in ("read", "simple"):
        for record in reads_simple(config):
            print(json.dumps(record), flush=True)
    elif args.mode == "stats":
        for record in reads_stats(config):
            print(json.dumps(record), flush=True)
    elif args.mode == "raw":
        for record in reads_raw(config):
            print(json.dumps(record), flush=True)
    elif args.mode == "table":
        for record in reads_table(config):
            print(json.dumps(record), flush=True)
    else:
        writes(config)
