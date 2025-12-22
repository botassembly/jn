#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.0"]
# ///
"""
Create sample Excel files for demo showcasing all parsing modes.
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.comments import Comment

# ============================================================================
# File 1: budget.xlsx - Simple clean table (demonstrates simple mode)
# ============================================================================

budget_data = [
    ["Month", "Category", "Description", "Amount"],
    ["January", "Engineering", "Software Licenses", 1200],
    ["January", "Marketing", "Ad Campaign", 3500],
    ["January", "Operations", "Office Rent", 4000],
    ["January", "HR", "Recruiting", 800],
    ["February", "Engineering", "Cloud Services", 2200],
    ["February", "Marketing", "Content Creation", 1500],
    ["February", "Operations", "Utilities", 450],
    ["February", "Sales", "CRM Software", 600],
    ["March", "Engineering", "Development Tools", 950],
    ["March", "Marketing", "Social Media Ads", 2800],
    ["March", "Operations", "Office Supplies", 320],
    ["March", "HR", "Training Programs", 1200],
]

wb = Workbook()
ws = wb.active
ws.title = "Budget 2024"

for row in budget_data:
    ws.append(row)

wb.save("budget.xlsx")
print("✓ Created budget.xlsx (simple table)")

# ============================================================================
# File 2: report.xlsx - Complex workbook (demonstrates stats, raw, table modes)
# ============================================================================

wb2 = Workbook()

# Sheet 1: Summary with title rows and merged cells
ws1 = wb2.active
ws1.title = "Summary"

# Title rows (will need to skip these)
ws1["A1"] = "Q4 2024 Sales Report"
ws1.merge_cells("A1:D1")
ws1["A2"] = "Generated: 2024-12-22"

# Empty row 3

# Headers in row 4
ws1["A4"] = "Region"
ws1["B4"] = "Q4 Total"
ws1["C4"] = "Target"
ws1["D4"] = "Variance"

# Data rows
ws1["A5"] = "North"
ws1["B5"] = 150000
ws1["C5"] = 140000
ws1["D5"] = "=B5-C5"

ws1["A6"] = "South"
ws1["B6"] = 120000
ws1["C6"] = 130000
ws1["D6"] = "=B6-C6"

ws1["A7"] = "East"
ws1["B7"] = 180000
ws1["C7"] = 175000
ws1["D7"] = "=B7-C7"

ws1["A8"] = "West"
ws1["B8"] = 95000
ws1["C8"] = 100000
ws1["D8"] = "=B8-C8"

# Total row
ws1["A9"] = "Total"
ws1["B9"] = "=SUM(B5:B8)"
ws1["C9"] = "=SUM(C5:C8)"
ws1["D9"] = "=SUM(D5:D8)"

# Add a comment
ws1["B5"].comment = Comment("Exceeded expectations!", "Sales Team")

# Sheet 2: Monthly breakdown (clean table)
ws2 = wb2.create_sheet("Monthly")
monthly_data = [
    ["Month", "North", "South", "East", "West"],
    ["October", 48000, 38000, 58000, 30000],
    ["November", 52000, 42000, 62000, 32000],
    ["December", 50000, 40000, 60000, 33000],
]
for row in monthly_data:
    ws2.append(row)

# Format as currency
for row in ws2.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5):
    for cell in row:
        cell.number_format = "$#,##0"

# Sheet 3: Notes (sparse data)
ws3 = wb2.create_sheet("Notes")
ws3["A1"] = "Important Notes"
ws3["B3"] = "Review Q1 targets"
ws3["C5"] = date(2024, 12, 15)
ws3["A8"] = "Contact: sales@example.com"

# Hide row 3 for demo
ws3.row_dimensions[3].hidden = True

wb2.save("report.xlsx")
print("✓ Created report.xlsx (multi-sheet with merged cells, formulas, comments)")

print("\nFiles ready for demo:")
print("  budget.xlsx - Simple table (simple mode)")
print("  report.xlsx - Complex workbook (stats, raw, table modes)")
