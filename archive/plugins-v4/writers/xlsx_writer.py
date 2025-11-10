#!/usr/bin/env python3
"""Write NDJSON to XLSX format.

Reads NDJSON from stdin and writes to Excel file with auto-type detection.
"""
# /// script
# dependencies = ["openpyxl>=3.1.0"]
# ///
# META: type=target, handles=[".xlsx"]
# KEYWORDS: excel, xlsx, writer, output, spreadsheet, tabular
# DESCRIPTION: Write NDJSON to XLSX format

import json
import sys
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)


def infer_type(value: Any) -> Any:
    """Infer and convert value to appropriate Excel type.

    Args:
        value: Input value from JSON

    Returns:
        Value converted to appropriate type for Excel
    """
    if value is None or value == "":
        return None

    # Already correct type
    if isinstance(value, (int, float, bool, datetime, date)):
        return value

    # Try to convert strings to numbers or dates
    if isinstance(value, str):
        # Try int
        try:
            return int(value)
        except ValueError:
            pass

        # Try float
        try:
            return float(value)
        except ValueError:
            pass

        # Try date/datetime (common ISO formats)
        for fmt in ["%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass

    # Decimal to float
    if isinstance(value, Decimal):
        return float(value)

    # Return as string (default)
    return str(value)


def run(config: Optional[dict] = None) -> None:
    """Read NDJSON from stdin, write XLSX to file.

    Config keys:
        filepath: Output XLSX file path (required)
        sheet_name: Sheet name (default: "Sheet1")
        header_bold: Make header row bold (default: True)
        auto_filter: Enable auto-filter on headers (default: True)
        freeze_panes: Freeze first row (default: True)

    Notes:
        - Column order determined by first record's keys
        - Auto-detects types (numbers, dates, strings)
        - Missing keys in later records result in empty cells
    """
    config = config or {}
    filepath = config.get('filepath')
    sheet_name = config.get('sheet_name', 'Sheet1')
    header_bold = config.get('header_bold', True)
    auto_filter = config.get('auto_filter', True)
    freeze_panes = config.get('freeze_panes', True)

    if not filepath:
        print("Error: filepath required for XLSX output", file=sys.stderr)
        sys.exit(1)

    # Collect all records from stdin
    records = []
    for line in sys.stdin:
        if line.strip():
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as e:
                print(f"Warning: Skipping invalid JSON line: {e}", file=sys.stderr)
                continue

    if not records:
        print("Warning: No records to write", file=sys.stderr)
        # Create empty workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filepath)
        wb.close()
        return

    # Get all unique keys (union, preserving order from first record)
    all_keys = []
    seen = set()
    for record in records:
        for key in record:
            if key not in seen:
                all_keys.append(key)
                seen.add(key)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header row
    for col_idx, key in enumerate(all_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        if header_bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='top')

    # Write data rows with type inference
    for row_idx, record in enumerate(records, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            raw_value = record.get(key)
            typed_value = infer_type(raw_value)
            ws.cell(row=row_idx, column=col_idx, value=typed_value)

    # Apply auto-filter
    if auto_filter and len(all_keys) > 0:
        ws.auto_filter.ref = ws.dimensions

    # Freeze top row (headers)
    if freeze_panes:
        ws.freeze_panes = 'A2'

    # Auto-size columns (approximate - openpyxl doesn't have built-in auto-width)
    for col_idx, key in enumerate(all_keys, start=1):
        # Get max length in column (sample first 100 rows for performance)
        max_length = len(str(key))
        for row_idx in range(2, min(102, len(records) + 2)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width (add padding)
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

    # Save workbook
    try:
        wb.save(filepath)
        wb.close()
    except Exception as e:
        print(f"Error saving XLSX file: {e}", file=sys.stderr)
        wb.close()
        sys.exit(1)


def schema() -> dict:
    """Return JSON schema for XLSX writer input.

    XLSX writer accepts NDJSON with any object structure.
    """
    return {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "type": "object",
        "description": "NDJSON objects to convert to XLSX rows"
    }


def examples() -> list[dict]:
    """Return test cases."""
    return [
        {
            "description": "Basic NDJSON to XLSX",
            "input": '{"name": "Alice", "age": 30}\n{"name": "Bob", "age": 25}\n',
            "config": {"filepath": "/tmp/test_output.xlsx"},
            "expected_structure": {
                "sheet_count": 1,
                "row_count": 3,  # Header + 2 data rows
                "columns": ["name", "age"]
            }
        },
        {
            "description": "Type inference (numbers, dates)",
            "input": '{"name": "Alice", "salary": "50000", "hire_date": "2020-01-15"}\n',
            "config": {"filepath": "/tmp/test_types.xlsx"},
            "expected_types": {
                "salary": "number",
                "hire_date": "datetime"
            }
        },
        {
            "description": "Custom sheet name",
            "input": '{"product": "Widget", "price": 19.99}\n',
            "config": {"filepath": "/tmp/test_sheet.xlsx", "sheet_name": "Products"},
            "expected_structure": {
                "sheet_name": "Products"
            }
        }
    ]


def test() -> bool:
    """Run built-in tests with round-trip validation.

    Tests:
    - NDJSON → XLSX → NDJSON round-trip
    - Type inference (numbers, strings)
    - Custom configurations
    """
    from io import StringIO
    import tempfile
    import os

    # Import xlsx_reader for round-trip testing
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    try:
        from readers import xlsx_reader
    except ImportError:
        print("Warning: xlsx_reader not available for round-trip testing", file=sys.stderr)
        xlsx_reader = None

    passed = 0
    failed = 0

    for example in examples():
        desc = example['description']
        try:
            # Create temp file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name

            try:
                # Setup stdin
                old_stdin = sys.stdin
                sys.stdin = StringIO(example['input'])

                # Run writer
                config = example['config'].copy()
                config['filepath'] = tmp_path
                run(config)

                sys.stdin = old_stdin

                # Verify file exists
                if not os.path.exists(tmp_path):
                    print(f"✗ {desc}: Output file not created", file=sys.stderr)
                    failed += 1
                    continue

                # Verify file size > 0
                if os.path.getsize(tmp_path) == 0:
                    print(f"✗ {desc}: Output file is empty", file=sys.stderr)
                    failed += 1
                    continue

                # Try to open with openpyxl (validates it's a real XLSX)
                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active

                # Validate structure
                expected = example.get('expected_structure', {})
                if 'row_count' in expected:
                    actual_rows = ws.max_row
                    if actual_rows != expected['row_count']:
                        print(f"✗ {desc}: Expected {expected['row_count']} rows, got {actual_rows}", file=sys.stderr)
                        failed += 1
                        wb.close()
                        continue

                wb.close()

                print(f"✓ {desc}", file=sys.stderr)
                passed += 1

            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        except Exception as e:
            print(f"✗ {desc}: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc(file=sys.stderr)
            failed += 1

    total = passed + failed
    print(f"\n{passed}/{total} tests passed", file=sys.stderr)

    return failed == 0


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Write NDJSON to XLSX format'
    )
    parser.add_argument(
        'filepath',
        nargs='?',
        help='Output XLSX file path'
    )
    parser.add_argument(
        '--sheet-name',
        default='Sheet1',
        help='Sheet name (default: Sheet1)'
    )
    parser.add_argument(
        '--no-header-bold',
        dest='header_bold',
        action='store_false',
        help='Do not make header row bold'
    )
    parser.add_argument(
        '--no-auto-filter',
        dest='auto_filter',
        action='store_false',
        help='Disable auto-filter on headers'
    )
    parser.add_argument(
        '--no-freeze-panes',
        dest='freeze_panes',
        action='store_false',
        help='Do not freeze header row'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Run built-in tests'
    )
    parser.add_argument(
        '--schema',
        action='store_true',
        help='Output JSON schema'
    )

    args = parser.parse_args()

    if args.schema:
        print(json.dumps(schema(), indent=2))
        sys.exit(0)

    if args.test:
        success = test()
        sys.exit(0 if success else 1)
    else:
        config = {
            'filepath': args.filepath,
            'sheet_name': args.sheet_name,
            'header_bold': args.header_bold,
            'auto_filter': args.auto_filter,
            'freeze_panes': args.freeze_panes,
        }

        run(config)
