"""CLI integration tests for XLSX plugin using jn cat command."""
import io
import json
import tempfile
from pathlib import Path

import openpyxl


def test_jn_cat_xlsx_file(invoke):
    """Test reading XLSX file using jn cat command."""
    # Create a test XLSX file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    # Write header and data
    sheet.cell(row=1, column=1, value="product")
    sheet.cell(row=1, column=2, value="price")
    sheet.cell(row=2, column=1, value="Apple")
    sheet.cell(row=2, column=2, value=1.50)
    sheet.cell(row=3, column=1, value="Banana")
    sheet.cell(row=3, column=2, value=0.75)

    # Save to temp file
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        workbook.save(f.name)
        temp_path = f.name

    try:
        # Test jn cat
        result = invoke(["cat", temp_path])
        assert result.exit_code == 0, f"Failed with: {result.output}"

        lines = [l for l in result.output.strip().split("\n") if l]
        assert len(lines) == 2

        record1 = json.loads(lines[0])
        assert record1["product"] == "Apple"
        assert record1["price"] == 1.5

        record2 = json.loads(lines[1])
        assert record2["product"] == "Banana"
        assert record2["price"] == 0.75
    finally:
        Path(temp_path).unlink()


def test_jn_cat_xlsx_then_put_csv(invoke):
    """Test converting XLSX to CSV using jn cat | jn put pipeline."""
    # Create a test XLSX file
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="name")
    sheet.cell(row=1, column=2, value="score")
    sheet.cell(row=2, column=1, value="Alice")
    sheet.cell(row=2, column=2, value=95)
    sheet.cell(row=3, column=1, value="Bob")
    sheet.cell(row=3, column=2, value=87)

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        workbook.save(f.name)
        xlsx_path = f.name

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        csv_path = f.name

    try:
        # Read XLSX
        result1 = invoke(["cat", xlsx_path])
        assert result1.exit_code == 0

        # Write to CSV
        result2 = invoke(["put", csv_path], input_data=result1.output)
        assert result2.exit_code == 0

        # Verify CSV content
        csv_content = Path(csv_path).read_text()
        lines = csv_content.strip().split("\n")
        assert lines[0] == "name,score"
        assert "Alice,95" in lines[1]
        assert "Bob,87" in lines[2]
    finally:
        Path(xlsx_path).unlink()
        Path(csv_path).unlink(missing_ok=True)


def test_jn_cat_xlsx_multisheet(invoke):
    """Test reading specific sheet from multi-sheet XLSX."""
    workbook = openpyxl.Workbook()

    # First sheet
    sheet1 = workbook.active
    sheet1.title = "Sales"
    sheet1.cell(row=1, column=1, value="item")
    sheet1.cell(row=2, column=1, value="Widget")

    # Second sheet
    sheet2 = workbook.create_sheet("Inventory")
    sheet2.cell(row=1, column=1, value="stock")
    sheet2.cell(row=2, column=1, value=100)

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        workbook.save(f.name)
        temp_path = f.name

    try:
        # Default sheet (first)
        result = invoke(["cat", temp_path])
        assert result.exit_code == 0
        record = json.loads(result.output.strip())
        assert record["item"] == "Widget"
    finally:
        Path(temp_path).unlink()


def test_jn_cat_xlsx_empty_rows(invoke):
    """Test that empty rows are skipped when reading XLSX."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Header
    sheet.cell(row=1, column=1, value="name")
    sheet.cell(row=1, column=2, value="value")

    # Data with empty row in middle
    sheet.cell(row=2, column=1, value="First")
    sheet.cell(row=2, column=2, value=10)
    # Row 3 is completely empty
    sheet.cell(row=4, column=1, value="Second")
    sheet.cell(row=4, column=2, value=20)

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        workbook.save(f.name)
        temp_path = f.name

    try:
        result = invoke(["cat", temp_path])
        assert result.exit_code == 0

        lines = [l for l in result.output.strip().split("\n") if l]
        assert len(lines) == 2  # Empty row should be skipped

        assert json.loads(lines[0])["name"] == "First"
        assert json.loads(lines[1])["name"] == "Second"
    finally:
        Path(temp_path).unlink()


def test_jn_put_xlsx(invoke):
    """Test writing XLSX file using jn put command."""
    ndjson = '{"city":"NYC","population":8336817}\n{"city":"LA","population":3979576}\n'

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        xlsx_path = f.name

    try:
        # Write XLSX
        result = invoke(["put", xlsx_path], input_data=ndjson)
        assert result.exit_code == 0, f"Failed with: {result.output}"

        # Verify XLSX content
        workbook = openpyxl.load_workbook(xlsx_path)
        sheet = workbook.active

        # Check header
        assert sheet.cell(row=1, column=1).value == "city"
        assert sheet.cell(row=1, column=2).value == "population"

        # Check data
        assert sheet.cell(row=2, column=1).value == "NYC"
        assert sheet.cell(row=2, column=2).value == 8336817
        assert sheet.cell(row=3, column=1).value == "LA"
        assert sheet.cell(row=3, column=2).value == 3979576
    finally:
        Path(xlsx_path).unlink(missing_ok=True)


def test_jn_run_csv_to_xlsx(invoke):
    """Test converting CSV to XLSX using jn run command."""
    csv_content = "country,capital\nFrance,Paris\nJapan,Tokyo\n"

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        f.write(csv_content)
        csv_path = f.name

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        xlsx_path = f.name

    try:
        # Convert CSV → XLSX
        result = invoke(["run", csv_path, xlsx_path])
        assert result.exit_code == 0, f"Failed with: {result.output}"

        # Verify XLSX content
        workbook = openpyxl.load_workbook(xlsx_path)
        sheet = workbook.active

        assert sheet.cell(row=1, column=1).value == "country"
        assert sheet.cell(row=1, column=2).value == "capital"
        assert sheet.cell(row=2, column=1).value == "France"
        assert sheet.cell(row=2, column=2).value == "Paris"
        assert sheet.cell(row=3, column=1).value == "Japan"
        assert sheet.cell(row=3, column=2).value == "Tokyo"
    finally:
        Path(csv_path).unlink()
        Path(xlsx_path).unlink(missing_ok=True)
