import io


def test_run_stdin_to_xlsx(invoke, tmp_path, sample_ndjson):
    out = tmp_path / "out.xlsx"
    res = invoke(["run", "-", str(out)], input_data=sample_ndjson)
    assert res.exit_code == 0, res.output

    # Validate XLSX
    import openpyxl

    wb = openpyxl.load_workbook(out)
    sheet = wb.active
    assert sheet.cell(row=1, column=1).value in ("name", "age")
    assert sheet.max_row >= 3
