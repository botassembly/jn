import io
import json


def test_plugin_call_xlsx_write(invoke):
    """Test writing XLSX via 'jn plugin call' command."""
    ndjson = '{"name":"Alice","age":30}\n{"name":"Bob","age":25}\n'

    # Call via jn plugin call command (tests binary output handling in service.py)
    result = invoke(["plugin", "call", "xlsx_", "--mode", "write"], input_data=ndjson)

    assert result.exit_code == 0, f"Failed: {result.output}"

    # Verify output is valid XLSX
    import openpyxl

    xlsx_bytes = result.stdout_bytes if hasattr(result, 'stdout_bytes') else result.output.encode('latin-1')
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active

    # Check header
    assert sheet.cell(row=1, column=1).value == "name"
    assert sheet.cell(row=1, column=2).value == "age"

    # Check data
    assert sheet.cell(row=2, column=1).value == "Alice"
    assert sheet.cell(row=2, column=2).value == 30
    assert sheet.cell(row=3, column=1).value == "Bob"
    assert sheet.cell(row=3, column=2).value == 25
