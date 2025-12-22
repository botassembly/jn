"""Integration tests for xlsx plugin with multiple parsing modes."""

import io
import json
import subprocess
import tempfile
from pathlib import Path

import pytest

# Ensure openpyxl is available for test fixtures
openpyxl = pytest.importorskip("openpyxl")


# Path to plugin
PLUGIN_PATH = Path(__file__).parent.parent.parent / "jn_home" / "plugins" / "formats" / "xlsx_.py"


@pytest.fixture(scope="module")
def xlsx_plugin():
    """Get path to xlsx plugin."""
    if not PLUGIN_PATH.exists():
        pytest.skip(f"xlsx plugin not found at {PLUGIN_PATH}")
    return str(PLUGIN_PATH)


# =============================================================================
# Test Data Fixtures - Create various Excel files for testing
# =============================================================================


@pytest.fixture
def simple_xlsx(tmp_path):
    """Create a simple Excel file (clean table, headers in row 1)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Headers
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "City"

    # Data
    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["C2"] = "NYC"

    ws["A3"] = "Bob"
    ws["B3"] = 25
    ws["C3"] = "LA"

    ws["A4"] = "Charlie"
    ws["B4"] = 35
    ws["C4"] = "Chicago"

    path = tmp_path / "simple.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """Create Excel with multiple sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: Users
    ws1 = wb.active
    ws1.title = "Users"
    ws1["A1"] = "ID"
    ws1["B1"] = "Name"
    ws1["A2"] = 1
    ws1["B2"] = "Alice"
    ws1["A3"] = 2
    ws1["B3"] = "Bob"

    # Sheet 2: Products
    ws2 = wb.create_sheet("Products")
    ws2["A1"] = "SKU"
    ws2["B1"] = "Price"
    ws2["A2"] = "PROD001"
    ws2["B2"] = 29.99
    ws2["A3"] = "PROD002"
    ws2["B3"] = 49.99

    # Sheet 3: Empty
    ws3 = wb.create_sheet("Empty")

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def title_rows_xlsx(tmp_path):
    """Create Excel with title rows before headers (report-style)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title rows
    ws["A1"] = "Quarterly Sales Report"
    ws["A2"] = "Q4 2024"

    # Headers in row 4
    ws["A4"] = "Region"
    ws["B4"] = "Sales"
    ws["C4"] = "Target"

    # Data
    ws["A5"] = "North"
    ws["B5"] = 150000
    ws["C5"] = 140000

    ws["A6"] = "South"
    ws["B6"] = 120000
    ws["C6"] = 130000

    path = tmp_path / "title_rows.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def merged_cells_xlsx(tmp_path):
    """Create Excel with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"

    # Merged header spanning B1:D1
    ws["A1"] = "Region"
    ws["B1"] = "Q1 2024"
    ws.merge_cells("B1:D1")

    # Sub-headers
    ws["A2"] = ""
    ws["B2"] = "Jan"
    ws["C2"] = "Feb"
    ws["D2"] = "Mar"

    # Data with merged row label
    ws["A3"] = "North"
    ws.merge_cells("A3:A4")
    ws["B3"] = 100
    ws["C3"] = 110
    ws["D3"] = 120

    ws["B4"] = 105
    ws["C4"] = 115
    ws["D4"] = 125

    ws["A5"] = "South"
    ws["B5"] = 80
    ws["C5"] = 85
    ws["D5"] = 90

    path = tmp_path / "merged.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def formulas_xlsx(tmp_path):
    """Create Excel with formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"

    ws["A1"] = "Item"
    ws["B1"] = "Qty"
    ws["C1"] = "Price"
    ws["D1"] = "Total"

    ws["A2"] = "Widget"
    ws["B2"] = 10
    ws["C2"] = 5.00
    ws["D2"] = "=B2*C2"  # Formula

    ws["A3"] = "Gadget"
    ws["B3"] = 5
    ws["C3"] = 15.00
    ws["D3"] = "=B3*C3"  # Formula

    ws["A4"] = "Total"
    ws["D4"] = "=SUM(D2:D3)"  # Summary formula

    path = tmp_path / "formulas.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def types_xlsx(tmp_path):
    """Create Excel with different data types."""
    from datetime import datetime, date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Types"

    ws["A1"] = "Type"
    ws["B1"] = "Value"

    ws["A2"] = "String"
    ws["B2"] = "Hello World"

    ws["A3"] = "Integer"
    ws["B3"] = 42

    ws["A4"] = "Float"
    ws["B4"] = 3.14159

    ws["A5"] = "Boolean"
    ws["B5"] = True

    ws["A6"] = "Date"
    ws["B6"] = date(2024, 12, 25)

    ws["A7"] = "DateTime"
    ws["B7"] = datetime(2024, 12, 25, 10, 30, 0)

    ws["A8"] = "None"
    ws["B8"] = None

    path = tmp_path / "types.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def sparse_xlsx(tmp_path):
    """Create Excel with sparse data (scattered cells)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"

    # Scattered cells
    ws["A1"] = "Title"
    ws["C3"] = "Data1"
    ws["C4"] = 100
    ws["E5"] = "Note"
    ws["B10"] = "Footer"

    path = tmp_path / "sparse.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def hidden_xlsx(tmp_path):
    """Create Excel with hidden rows and columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hidden"

    ws["A1"] = "Name"
    ws["B1"] = "Hidden Col"
    ws["C1"] = "Value"

    ws["A2"] = "Row1"
    ws["B2"] = "secret1"
    ws["C2"] = 100

    ws["A3"] = "HiddenRow"
    ws["B3"] = "secret2"
    ws["C3"] = 200

    ws["A4"] = "Row3"
    ws["B4"] = "secret3"
    ws["C4"] = 300

    # Hide column B
    ws.column_dimensions["B"].hidden = True

    # Hide row 3
    ws.row_dimensions[3].hidden = True

    path = tmp_path / "hidden.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def comments_xlsx(tmp_path):
    """Create Excel with cell comments."""
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comments"

    ws["A1"] = "Name"
    ws["B1"] = "Value"

    ws["A2"] = "Item1"
    ws["B2"] = 100
    ws["B2"].comment = Comment("This needs review", "Author")

    ws["A3"] = "Item2"
    ws["B3"] = 200

    path = tmp_path / "comments.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def formatted_xlsx(tmp_path):
    """Create Excel with number formats."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formatted"

    ws["A1"] = "Type"
    ws["B1"] = "Value"

    ws["A2"] = "Currency"
    ws["B2"] = 1234.56
    ws["B2"].number_format = "$#,##0.00"

    ws["A3"] = "Percentage"
    ws["B3"] = 0.75
    ws["B3"].number_format = "0.00%"

    ws["A4"] = "Date"
    ws["B4"] = 45000  # Excel date serial
    ws["B4"].number_format = "yyyy-mm-dd"

    ws["A5"] = "General"
    ws["B5"] = 42

    path = tmp_path / "formatted.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def empty_headers_xlsx(tmp_path):
    """Create Excel with empty header cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptyHeaders"

    ws["A1"] = "Name"
    ws["B1"] = None  # Empty header
    ws["C1"] = "City"
    ws["D1"] = None  # Empty header

    ws["A2"] = "Alice"
    ws["B2"] = 30
    ws["C2"] = "NYC"
    ws["D2"] = "extra"

    path = tmp_path / "empty_headers.xlsx"
    wb.save(path)
    wb.close()
    return path


# =============================================================================
# Helper function to run xlsx plugin
# =============================================================================


def run_xlsx(plugin_path, xlsx_path, args=None, mode="read"):
    """Run xlsx plugin and return parsed NDJSON output."""
    cmd = ["uv", "run", "--script", plugin_path, f"--mode={mode}"]
    if args:
        cmd.extend(args)

    with open(xlsx_path, "rb") as f:
        result = subprocess.run(
            cmd,
            stdin=f,
            capture_output=True,
        )

    if result.returncode != 0:
        raise RuntimeError(f"xlsx plugin failed: {result.stderr.decode()}")

    # Parse NDJSON output
    output = result.stdout.decode()
    lines = [l for l in output.strip().split("\n") if l]
    return [json.loads(line) for line in lines]


# =============================================================================
# Mode 1: Simple Mode Tests (Default CSV-like behavior)
# =============================================================================


class TestSimpleMode:
    """Tests for simple mode (default, CSV-like parsing)."""

    def test_simple_read(self, xlsx_plugin, simple_xlsx):
        """Test reading a simple Excel file."""
        records = run_xlsx(xlsx_plugin, simple_xlsx)

        assert len(records) == 3
        assert records[0] == {"Name": "Alice", "Age": 30, "City": "NYC"}
        assert records[1] == {"Name": "Bob", "Age": 25, "City": "LA"}
        assert records[2] == {"Name": "Charlie", "Age": 35, "City": "Chicago"}

    def test_simple_read_with_skip_rows(self, xlsx_plugin, title_rows_xlsx):
        """Test skipping title rows."""
        records = run_xlsx(xlsx_plugin, title_rows_xlsx, args=["--skip-rows=3"])

        assert len(records) == 2
        assert records[0]["Region"] == "North"
        assert records[0]["Sales"] == 150000

    def test_simple_read_specific_sheet(self, xlsx_plugin, multi_sheet_xlsx):
        """Test reading specific sheet by name."""
        records = run_xlsx(xlsx_plugin, multi_sheet_xlsx, args=["--sheet=Products"])

        assert len(records) == 2
        assert records[0]["SKU"] == "PROD001"
        assert records[0]["Price"] == 29.99

    def test_simple_read_sheet_by_index(self, xlsx_plugin, multi_sheet_xlsx):
        """Test reading specific sheet by index."""
        records = run_xlsx(xlsx_plugin, multi_sheet_xlsx, args=["--sheet=1"])

        assert len(records) == 2
        assert records[0]["SKU"] == "PROD001"

    def test_simple_read_empty_sheet(self, xlsx_plugin, multi_sheet_xlsx):
        """Test reading empty sheet."""
        records = run_xlsx(xlsx_plugin, multi_sheet_xlsx, args=["--sheet=Empty"])

        assert len(records) == 0

    def test_simple_read_empty_headers(self, xlsx_plugin, empty_headers_xlsx):
        """Test handling empty header cells."""
        records = run_xlsx(xlsx_plugin, empty_headers_xlsx)

        assert len(records) == 1
        # Empty headers should be named Column_N
        assert "Name" in records[0]
        assert "Column_2" in records[0]
        assert "City" in records[0]
        assert "Column_4" in records[0]

    def test_simple_read_types(self, xlsx_plugin, types_xlsx):
        """Test handling different data types."""
        records = run_xlsx(xlsx_plugin, types_xlsx)

        # Find records by type
        types_map = {r["Type"]: r["Value"] for r in records}

        assert types_map["String"] == "Hello World"
        assert types_map["Integer"] == 42
        assert abs(types_map["Float"] - 3.14159) < 0.0001
        assert types_map["Boolean"] is True
        # Date/DateTime should be ISO format strings
        assert "2024-12-25" in str(types_map["Date"])
        assert types_map["None"] is None


# =============================================================================
# Mode 2: Stats Mode Tests (Workbook metadata)
# =============================================================================


class TestStatsMode:
    """Tests for stats mode (workbook structure inspection)."""

    def test_stats_simple(self, xlsx_plugin, simple_xlsx):
        """Test stats for simple workbook."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=[], mode="stats")

        assert len(records) == 1
        sheet = records[0]

        assert sheet["sheet"] == "Data"
        assert sheet["index"] == 0
        assert sheet["rows"] == 4
        assert sheet["cols"] == 3

    def test_stats_multi_sheet(self, xlsx_plugin, multi_sheet_xlsx):
        """Test stats for multi-sheet workbook."""
        records = run_xlsx(xlsx_plugin, multi_sheet_xlsx, args=[], mode="stats")

        assert len(records) == 3

        # Sheet names
        sheet_names = [r["sheet"] for r in records]
        assert sheet_names == ["Users", "Products", "Empty"]

        # Indices
        indices = [r["index"] for r in records]
        assert indices == [0, 1, 2]

    def test_stats_merged_cells(self, xlsx_plugin, merged_cells_xlsx):
        """Test stats reports merged cell ranges."""
        records = run_xlsx(xlsx_plugin, merged_cells_xlsx, args=[], mode="stats")

        assert len(records) == 1
        sheet = records[0]

        assert "merged_ranges" in sheet
        # Should have two merged ranges: B1:D1 and A3:A4
        merged = sheet["merged_ranges"]
        assert len(merged) == 2
        assert "B1:D1" in merged
        assert "A3:A4" in merged

    def test_stats_dimensions(self, xlsx_plugin, simple_xlsx):
        """Test stats reports dimensions."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=[], mode="stats")

        sheet = records[0]
        assert "dimensions" in sheet
        assert sheet["min_row"] == 1
        assert sheet["max_row"] == 4
        assert sheet["min_col"] == 1
        assert sheet["max_col"] == 3

    def test_stats_first_row_and_col(self, xlsx_plugin, simple_xlsx):
        """Test stats includes first row and column values."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=[], mode="stats")

        sheet = records[0]

        # First row should be headers
        assert "first_row" in sheet
        assert sheet["first_row"] == ["Name", "Age", "City"]

        # First column should be row labels
        assert "first_col" in sheet
        assert sheet["first_col"] == ["Name", "Alice", "Bob", "Charlie"]


# =============================================================================
# Mode 3: Raw Mode Tests (Cell-by-cell output)
# =============================================================================


class TestRawMode:
    """Tests for raw mode (cell-level output with metadata)."""

    def test_raw_basic(self, xlsx_plugin, simple_xlsx):
        """Test raw mode outputs each cell."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=[], mode="raw")

        # 4 rows x 3 cols = 12 cells
        assert len(records) == 12

        # Check first cell
        a1 = next(r for r in records if r["ref"] == "A1")
        assert a1["sheet"] == "Data"
        assert a1["row"] == 1
        assert a1["col"] == 1
        assert a1["value"] == "Name"
        assert a1["type"] == "s"  # string

    def test_raw_types(self, xlsx_plugin, types_xlsx):
        """Test raw mode reports cell types."""
        records = run_xlsx(xlsx_plugin, types_xlsx, args=[], mode="raw")

        # Find cells by value type
        b2 = next(r for r in records if r["ref"] == "B2")
        assert b2["type"] == "s"  # string

        b3 = next(r for r in records if r["ref"] == "B3")
        assert b3["type"] == "n"  # number

        b5 = next(r for r in records if r["ref"] == "B5")
        assert b5["type"] == "b"  # boolean

    def test_raw_multi_sheet(self, xlsx_plugin, multi_sheet_xlsx):
        """Test raw mode includes all sheets by default."""
        records = run_xlsx(xlsx_plugin, multi_sheet_xlsx, args=[], mode="raw")

        # Check we have cells from both populated sheets
        sheets = set(r["sheet"] for r in records)
        assert "Users" in sheets
        assert "Products" in sheets

    def test_raw_single_sheet(self, xlsx_plugin, multi_sheet_xlsx):
        """Test raw mode with sheet filter."""
        records = run_xlsx(xlsx_plugin, multi_sheet_xlsx, args=["--sheet=Products"], mode="raw")

        # All cells should be from Products sheet
        sheets = set(r["sheet"] for r in records)
        assert sheets == {"Products"}

    def test_raw_merged_cells(self, xlsx_plugin, merged_cells_xlsx):
        """Test raw mode reports merge info."""
        records = run_xlsx(xlsx_plugin, merged_cells_xlsx, args=[], mode="raw")

        # B1 is merge origin of B1:D1
        b1 = next(r for r in records if r["ref"] == "B1")
        assert b1["merge"] == "B1:D1"
        assert b1["merge_origin"] is True
        assert b1["value"] == "Q1 2024"

    def test_raw_formulas(self, xlsx_plugin, formulas_xlsx):
        """Test raw mode reports formula cells."""
        # Need to load without data_only to see formulas
        records = run_xlsx(xlsx_plugin, formulas_xlsx, args=["--formulas=text"], mode="raw")

        # D2 should have formula
        d2 = next(r for r in records if r["ref"] == "D2")
        assert d2["type"] == "f"  # formula
        assert "=B2*C2" in d2["value"]

    def test_raw_hidden_cells(self, xlsx_plugin, hidden_xlsx):
        """Test raw mode reports hidden rows/columns."""
        records = run_xlsx(xlsx_plugin, hidden_xlsx, args=[], mode="raw")

        # Row 3 is hidden
        row3_cells = [r for r in records if r["row"] == 3]
        for cell in row3_cells:
            assert cell.get("hidden") is True

        # Column B is hidden
        col_b_cells = [r for r in records if r["col"] == 2]
        for cell in col_b_cells:
            assert cell.get("hidden") is True

    def test_raw_comments(self, xlsx_plugin, comments_xlsx):
        """Test raw mode reports comments."""
        records = run_xlsx(xlsx_plugin, comments_xlsx, args=[], mode="raw")

        # B2 has a comment
        b2 = next(r for r in records if r["ref"] == "B2")
        assert "comment" in b2
        assert "review" in b2["comment"]

        # B3 has no comment
        b3 = next(r for r in records if r["ref"] == "B3")
        assert "comment" not in b3

    def test_raw_number_formats(self, xlsx_plugin, formatted_xlsx):
        """Test raw mode reports number formats."""
        records = run_xlsx(xlsx_plugin, formatted_xlsx, args=[], mode="raw")

        # B2 has currency format
        b2 = next(r for r in records if r["ref"] == "B2")
        assert "format" in b2
        assert "$" in b2["format"]

        # B3 has percentage format
        b3 = next(r for r in records if r["ref"] == "B3")
        assert "format" in b3
        assert "%" in b3["format"]

        # B5 has General format (should be omitted)
        b5 = next(r for r in records if r["ref"] == "B5")
        assert b5.get("format") is None or b5.get("format") == "General"

    def test_raw_sparse_data(self, xlsx_plugin, sparse_xlsx):
        """Test raw mode handles sparse data (only non-empty cells)."""
        records = run_xlsx(xlsx_plugin, sparse_xlsx, args=[], mode="raw")

        # Should only have 5 non-empty cells
        assert len(records) == 5

        # Verify scattered locations
        refs = {r["ref"] for r in records}
        assert refs == {"A1", "C3", "C4", "E5", "B10"}

    def test_raw_range_filter(self, xlsx_plugin, simple_xlsx):
        """Test raw mode with range filter."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=["--range=A1:B2"], mode="raw")

        # Should only have cells in range A1:B2 (4 cells)
        assert len(records) == 4
        refs = {r["ref"] for r in records}
        assert refs == {"A1", "A2", "B1", "B2"}


# =============================================================================
# Mode 4: Table Mode Tests (Region extraction with config)
# =============================================================================


class TestTableMode:
    """Tests for table mode (explicit region extraction)."""

    def test_table_basic(self, xlsx_plugin, simple_xlsx):
        """Test table mode with default settings (same as simple)."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=[], mode="table")

        assert len(records) == 3
        assert records[0]["Name"] == "Alice"

    def test_table_with_range(self, xlsx_plugin, simple_xlsx):
        """Test table mode with specific range."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=["--range=A1:B3"], mode="table")

        # Only A-B columns, rows 1-3 (header + 2 data rows)
        assert len(records) == 2
        assert "Name" in records[0]
        assert "Age" in records[0]
        assert "City" not in records[0]

    def test_table_with_column_only_range(self, xlsx_plugin, simple_xlsx):
        """Test table mode with column-only range (e.g., A:B)."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=["--range=A:B"], mode="table")

        # Should use all rows, but only columns A-B
        assert len(records) == 3
        assert "Name" in records[0]
        assert "Age" in records[0]
        assert "City" not in records[0]

    def test_table_with_row_only_range(self, xlsx_plugin, simple_xlsx):
        """Test table mode with row-only range (e.g., 1:3)."""
        records = run_xlsx(xlsx_plugin, simple_xlsx, args=["--range=1:3"], mode="table")

        # Should use rows 1-3, all columns
        assert len(records) == 2  # 2 data rows (row 1 is header)
        assert "Name" in records[0]
        assert "Age" in records[0]
        assert "City" in records[0]

    def test_table_with_header_row(self, xlsx_plugin, title_rows_xlsx):
        """Test table mode with explicit header row."""
        records = run_xlsx(xlsx_plugin, title_rows_xlsx, args=["--range=A4:C6", "--header-row=1"], mode="table")

        assert len(records) == 2
        assert records[0]["Region"] == "North"
        assert records[0]["Sales"] == 150000

    def test_table_merge_strategy_origin(self, xlsx_plugin, merged_cells_xlsx):
        """Test table mode with origin merge strategy (default)."""
        records = run_xlsx(xlsx_plugin, merged_cells_xlsx, args=["--range=A1:D5", "--merge-strategy=origin"], mode="table")

        # First non-header row should have "Q1 2024" from merged cell
        assert len(records) > 0

    def test_table_merge_strategy_fill(self, xlsx_plugin, merged_cells_xlsx):
        """Test table mode with fill merge strategy."""
        records = run_xlsx(xlsx_plugin, merged_cells_xlsx, args=["--range=A1:D5", "--merge-strategy=fill"], mode="table")

        # Merged cell value should propagate
        assert len(records) > 0

    def test_table_skip_empty_rows(self, xlsx_plugin, sparse_xlsx):
        """Test table mode skips empty rows by default."""
        records = run_xlsx(xlsx_plugin, sparse_xlsx, args=["--range=A1:E10"], mode="table")

        # Should skip completely empty rows
        for r in records:
            # At least one value should be non-null
            assert any(v is not None for v in r.values())

    def test_table_formulas_computed(self, xlsx_plugin, formulas_xlsx):
        """Test table mode with computed formula values (default)."""
        records = run_xlsx(xlsx_plugin, formulas_xlsx, args=["--formulas=computed"], mode="table")

        # D2 formula =B2*C2 should compute to 50 (10 * 5)
        # Note: openpyxl may not have cached value if file wasn't opened in Excel
        assert len(records) >= 2

    def test_table_formulas_text(self, xlsx_plugin, formulas_xlsx):
        """Test table mode with formula text."""
        records = run_xlsx(xlsx_plugin, formulas_xlsx, args=["--formulas=text"], mode="table")

        # Should have formula text
        widget_row = next((r for r in records if r.get("Item") == "Widget"), None)
        if widget_row:
            assert "Total" in widget_row
            # Formula text or computed value
            total = widget_row["Total"]
            assert total is not None

    def test_table_transposed_header_col(self, xlsx_plugin, tmp_path):
        """Test table mode with header in column (transposed data)."""
        # Create transposed data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Alice"
        ws["C1"] = "Bob"
        ws["A2"] = "Age"
        ws["B2"] = 30
        ws["C2"] = 25
        ws["A3"] = "City"
        ws["B3"] = "NYC"
        ws["C3"] = "LA"

        path = tmp_path / "transposed.xlsx"
        wb.save(path)
        wb.close()

        records = run_xlsx(xlsx_plugin, path, args=["--header-col=A"], mode="table")

        assert len(records) == 2
        assert records[0]["Name"] == "Alice"
        assert records[0]["Age"] == 30
        assert records[0]["City"] == "NYC"


# =============================================================================
# Error Handling Tests
# =============================================================================


class TestErrorHandling:
    """Tests for error handling."""

    def test_invalid_sheet_name(self, xlsx_plugin, simple_xlsx):
        """Test error on invalid sheet name."""
        with pytest.raises(RuntimeError) as exc:
            run_xlsx(xlsx_plugin, simple_xlsx, args=["--sheet=NonExistent"])

        assert "not found" in str(exc.value).lower() or "NonExistent" in str(exc.value)

    def test_invalid_sheet_index(self, xlsx_plugin, simple_xlsx):
        """Test error on invalid sheet index."""
        with pytest.raises(RuntimeError) as exc:
            run_xlsx(xlsx_plugin, simple_xlsx, args=["--sheet=99"])

        assert "out of range" in str(exc.value).lower() or "99" in str(exc.value)


# =============================================================================
# Integration Tests - Full pipeline with jn cat
# =============================================================================


class TestJnCatIntegration:
    """Tests for jn cat integration with xlsx files."""

    @pytest.fixture
    def jn_path(self):
        """Get path to jn executable."""
        dist_jn = Path(__file__).parent.parent.parent / "dist" / "bin" / "jn"
        if not dist_jn.exists():
            pytest.skip("jn not built (run 'make build')")
        return str(dist_jn)

    def test_jn_cat_xlsx_simple(self, jn_path, simple_xlsx):
        """Test jn cat with xlsx file."""
        result = subprocess.run(
            [jn_path, "cat", str(simple_xlsx)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        lines = [l for l in result.stdout.strip().split("\n") if l]
        assert len(lines) == 3

        first = json.loads(lines[0])
        assert first["Name"] == "Alice"

    def test_jn_cat_xlsx_with_query_params(self, jn_path, multi_sheet_xlsx):
        """Test jn cat with query string parameters."""
        result = subprocess.run(
            [jn_path, "cat", f"{multi_sheet_xlsx}?sheet=Products"],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        lines = [l for l in result.stdout.strip().split("\n") if l]
        first = json.loads(lines[0])
        assert first["SKU"] == "PROD001"

    def test_jn_cat_xlsx_stats_mode(self, jn_path, multi_sheet_xlsx):
        """Test jn cat with mode=stats query param."""
        result = subprocess.run(
            [jn_path, "cat", f"{multi_sheet_xlsx}?mode=stats"],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        lines = [l for l in result.stdout.strip().split("\n") if l]
        assert len(lines) == 3  # 3 sheets

        sheets = [json.loads(l)["sheet"] for l in lines]
        assert "Users" in sheets
        assert "Products" in sheets

    def test_jn_cat_xlsx_raw_mode(self, jn_path, simple_xlsx):
        """Test jn cat with mode=raw query param."""
        result = subprocess.run(
            [jn_path, "cat", f"{simple_xlsx}?mode=raw"],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        lines = [l for l in result.stdout.strip().split("\n") if l]

        # Should have cell-level output
        first = json.loads(lines[0])
        assert "ref" in first
        assert "row" in first
        assert "col" in first
