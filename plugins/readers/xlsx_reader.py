#!/usr/bin/env python3
"""Read XLSX/XLSM files and output NDJSON.

Supports sheet selection, formula evaluation, and merged cell handling.
"""
# /// script
# dependencies = ["openpyxl>=3.1.0"]
# ///
# META: type=source, handles=[".xlsx", ".xlsm"], streaming=true
# KEYWORDS: excel, xlsx, xlsm, spreadsheet, data, parsing, tabular
# DESCRIPTION: Read Excel files and output NDJSON

import json
import sys
from io import BytesIO
from typing import Iterator, Optional, Union

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)


def run(config: Optional[dict] = None) -> Iterator[dict]:
    """Read XLSX from stdin or file, yield records as dicts.

    Config keys:
        sheet: Sheet name (str) or index (int, 0-based). Default: 0 (first sheet)
        filepath: Path to XLSX file (if not reading from stdin)
        data_only: Evaluate formulas to values (default: True)
        skip_rows: Number of rows to skip before header (default: 0)
        max_rows: Maximum number of rows to read (default: None = all)

    Yields:
        Dict per Excel row with column headers as keys
    """
    config = config or {}
    sheet_identifier = config.get('sheet', 0)
    filepath = config.get('filepath')
    data_only = config.get('data_only', True)
    skip_rows = config.get('skip_rows', 0)
    max_rows = config.get('max_rows')

    # Load workbook from file or stdin
    try:
        if filepath:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=data_only)
        else:
            # Read from stdin (binary)
            data = sys.stdin.buffer.read()
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=data_only)
    except InvalidFileException as e:
        print(f"Error: Invalid XLSX file: {e}", file=sys.stderr)
        return
    except Exception as e:
        print(f"Error loading workbook: {e}", file=sys.stderr)
        return

    # Select sheet
    try:
        if isinstance(sheet_identifier, int):
            ws = wb.worksheets[sheet_identifier]
        else:
            ws = wb[sheet_identifier]
    except (IndexError, KeyError) as e:
        available = [ws.title for ws in wb.worksheets]
        print(f"Error: Sheet '{sheet_identifier}' not found. Available sheets: {available}", file=sys.stderr)
        wb.close()
        return

    # Get rows iterator
    rows = ws.iter_rows(values_only=True)

    # Skip initial rows if requested
    for _ in range(skip_rows):
        try:
            next(rows)
        except StopIteration:
            wb.close()
            return

    # First row after skip = headers
    try:
        headers = next(rows)
    except StopIteration:
        wb.close()
        return

    # Clean headers: convert to strings, handle None, strip whitespace
    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(headers)
    ]

    # Yield data rows
    row_count = 0
    for row in rows:
        # Stop if max_rows reached
        if max_rows is not None and row_count >= max_rows:
            break

        # Skip completely empty rows
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue

        # Create dict with headers
        # Handle rows shorter than header row by padding with None
        record = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            # Convert None to empty string for consistency
            record[header] = value if value is not None else ""

        yield record
        row_count += 1

    wb.close()


def schema() -> dict:
    """Return JSON schema for XLSX output.

    XLSX reader outputs records as objects with mixed-type values.
    Field names come from the first row (headers).
    """
    return {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "type": "object",
        "description": "XLSX row as key-value pairs",
        "patternProperties": {
            ".*": {
                "type": ["string", "number", "boolean", "null"],
                "description": "Cell values can be strings, numbers, booleans, or null"
            }
        }
    }


def examples() -> list[dict]:
    """Return test cases for this plugin.

    Note: These are inline examples. Real-world testing uses actual XLSX files
    from public URLs (see test() function).
    """
    return [
        {
            "description": "Basic XLSX with headers and data",
            "note": "This is a conceptual example. Actual tests use real XLSX files.",
            "expected_structure": {
                "columns": ["name", "age", "city"],
                "row_count": 2,
                "sample_row": {"name": "Alice", "age": 30, "city": "NYC"}
            }
        },
        {
            "description": "XLSX with multiple sheets",
            "config": {"sheet": "Sheet2"},
            "note": "Sheet selection by name"
        },
        {
            "description": "XLSX with formulas",
            "config": {"data_only": True},
            "note": "Formulas evaluated to values"
        }
    ]


def test() -> bool:
    """Run outside-in tests with real public XLSX files.

    Tests multiple scenarios:
    - Small template files
    - Public S3 buckets via HTTPS
    - GitHub raw files
    - Different XLSX structures

    Returns:
        True if all tests pass
    """
    import subprocess
    import tempfile
    import os

    passed = 0
    failed = 0

    # Test cases with real public URLs
    test_urls = [
        {
            "description": "GitHub HuBMAP sample template (small, stable)",
            "url": "https://raw.githubusercontent.com/hubmapconsortium/dataset-metadata-spreadsheet/main/sample-section/latest/sample-section.xlsx",
            "min_records": 1,  # Should have at least 1 data row
        },
        {
            "description": "GitHub EBI EVA submission template",
            "url": "https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/main/eva_sub_cli/etc/EVA_Submission_Example.xlsx",
            "min_records": 1,
        },
        {
            "description": "GitHub COEF test.xlsx",
            "url": "https://raw.githubusercontent.com/Russel88/COEF/master/ExampleData/test.xlsx",
            "min_records": 1,
        }
    ]

    for test_case in test_urls:
        desc = test_case['description']
        url = test_case['url']
        min_records = test_case['min_records']

        try:
            # Download XLSX file
            print(f"Testing: {desc}", file=sys.stderr)
            download = subprocess.run(
                ["curl", "-sL", url],
                capture_output=True,
                timeout=30
            )

            if download.returncode != 0:
                print(f"  ✗ Failed to download: {download.stderr.decode()}", file=sys.stderr)
                failed += 1
                continue

            # Write to temp file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp.write(download.stdout)
                tmp_path = tmp.name

            try:
                # Parse XLSX
                config = {'filepath': tmp_path}
                records = list(run(config))

                # Validate
                if len(records) >= min_records:
                    # Check that records are dicts
                    if all(isinstance(r, dict) for r in records):
                        # Check that records have keys (columns)
                        if all(len(r) > 0 for r in records):
                            print(f"  ✓ {desc} ({len(records)} records)", file=sys.stderr)
                            passed += 1
                        else:
                            print(f"  ✗ {desc}: Records have no columns", file=sys.stderr)
                            failed += 1
                    else:
                        print(f"  ✗ {desc}: Invalid record format", file=sys.stderr)
                        failed += 1
                else:
                    print(f"  ✗ {desc}: Expected at least {min_records} records, got {len(records)}", file=sys.stderr)
                    failed += 1

            finally:
                os.unlink(tmp_path)

        except subprocess.TimeoutExpired:
            print(f"  ✗ {desc}: Download timeout", file=sys.stderr)
            failed += 1
        except Exception as e:
            print(f"  ✗ {desc}: {e}", file=sys.stderr)
            failed += 1

    total = passed + failed
    print(f"\n{passed}/{total} tests passed", file=sys.stderr)

    return failed == 0


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Read XLSX/XLSM and output NDJSON'
    )
    parser.add_argument(
        'filepath',
        nargs='?',
        help='Path to XLSX file (if not reading from stdin)'
    )
    parser.add_argument(
        '--sheet',
        help='Sheet name or index (0-based integer)'
    )
    parser.add_argument(
        '--data-only',
        action='store_true',
        default=True,
        help='Evaluate formulas to values (default: True)'
    )
    parser.add_argument(
        '--skip-rows',
        type=int,
        default=0,
        help='Number of rows to skip before header'
    )
    parser.add_argument(
        '--max-rows',
        type=int,
        help='Maximum number of rows to read'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Run outside-in tests with real public XLSX files'
    )
    parser.add_argument(
        '--schema',
        action='store_true',
        help='Output JSON schema'
    )

    args = parser.parse_args()

    if args.schema:
        print(json.dumps(schema(), indent=2))
        sys.exit(0)

    if args.test:
        success = test()
        sys.exit(0 if success else 1)
    else:
        # Normal operation: read XLSX, write NDJSON to stdout
        config = {
            'filepath': args.filepath,
            'skip_rows': args.skip_rows,
            'data_only': args.data_only,
        }

        # Handle sheet parameter (could be name or index)
        if args.sheet:
            try:
                config['sheet'] = int(args.sheet)
            except ValueError:
                config['sheet'] = args.sheet

        if args.max_rows:
            config['max_rows'] = args.max_rows

        for record in run(config):
            print(json.dumps(record, default=str), flush=True)
